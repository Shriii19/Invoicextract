"""
Excel exporter for invoice data
"""

import logging
from pathlib import Path
from typing import List, Union

from invoicextract.exporters.base_exporter import BaseExporter
from invoicextract.models.invoice import Invoice

logger = logging.getLogger(__name__)


class ExcelExporter(BaseExporter):
    """
    Exports invoice data to Excel format (.xlsx)

    Note: Requires openpyxl or xlsxwriter library
    """

    def export(self, invoices: List[Invoice], output_path: Union[str, Path]):
        """
        Export invoices to Excel file

        Args:
            invoices: List of Invoice objects to export
            output_path: Path to output Excel file
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        logger.info(f"Exporting {len(invoices)} invoices to Excel: {output_path}")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices"

            # Headers
            headers = [
                "Invoice #",
                "Vendor",
                "Customer",
                "Invoice Date",
                "Due Date",
                "Subtotal",
                "Tax",
                "Total",
                "Currency",
                "Payment Terms",
                "Valid",
                "Source File",
            ]

            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_idx, invoice in enumerate(invoices, 2):
                ws.cell(row=row_idx, column=1, value=invoice.invoice_number)
                ws.cell(row=row_idx, column=2, value=invoice.vendor_name)
                ws.cell(row=row_idx, column=3, value=invoice.customer_name)
                ws.cell(row=row_idx, column=4, value=invoice.invoice_date)
                ws.cell(row=row_idx, column=5, value=invoice.due_date)
                ws.cell(row=row_idx, column=6, value=invoice.subtotal)
                ws.cell(row=row_idx, column=7, value=invoice.tax_amount)
                ws.cell(row=row_idx, column=8, value=invoice.total_amount)
                ws.cell(row=row_idx, column=9, value=invoice.currency)
                ws.cell(row=row_idx, column=10, value=invoice.payment_terms)
                ws.cell(row=row_idx, column=11, value="Yes" if invoice.is_valid else "No")
                ws.cell(row=row_idx, column=12, value=invoice.source_file)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save workbook
            wb.save(output_path)
            logger.info(f"Successfully exported to {output_path}")

        except ImportError:
            logger.error("openpyxl library not found. Please install: pip install openpyxl")
            raise ImportError(
                "Excel export requires openpyxl library. " "Install with: pip install openpyxl"
            )
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise
